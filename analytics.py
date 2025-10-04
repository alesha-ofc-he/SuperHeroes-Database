import pandas as pd
import matplotlib.pyplot as plt
import os
from sqlalchemy import create_engine
import plotly.express as px
from openpyxl.formatting.rule import ColorScaleRule, FormulaRule
from openpyxl.styles import Font, Color, PatternFill
import sys
import time
import numpy as np 

# --- CONFIGURATION ---

# DB_URL is fetched from the Docker environment variable specified in docker-compose.yml
class Config:
    DB_URL = os.environ.get("DB_URL") 
    if not DB_URL:
        # Fallback for local run (using the exposed port 5433)
        DB_URL = "postgresql+psycopg2://postgres:1234@localhost:5433/superheroes" 

# Create the SQLAlchemy engine
try:
    engine = create_engine(Config.DB_URL)
except Exception as e:
    # If connection fails instantly (common when Docker setup is incomplete)
    print(f"Error creating SQLAlchemy engine: {e}", file=sys.stderr)
    sys.exit(1)


# Ensure output directories exist
os.makedirs('charts', exist_ok=True)
os.makedirs('exports', exist_ok=True)

# --- UTILITY FUNCTIONS ---

def run_query(sql_query, title):
    """Executes an SQL query and returns a DataFrame, printing a brief report."""
    
    # Simple retry mechanism to handle quick startup failures
    retries = 5
    for i in range(retries):
        try:
            df = pd.read_sql(sql_query, engine)
            print(f"\n--- Data Report: {title} ---")
            print(f"Rows fetched: {len(df)}")
            return df
        except Exception as e:
            if i < retries - 1:
                print(f"Connection attempt {i+1}/{retries} failed. Retrying in 5 seconds...")
                time.sleep(5)
            else:
                # Fatal error, the application cannot continue
                print(f"FATAL ERROR: Failed to connect to the database after {retries} attempts: {e}", file=sys.stderr)
                sys.exit(1)


def generate_chart(df, chart_type, title, filename, x_label=None, y_label=None, x_col=None, y_col=None, **kwargs):
    """Generates a Matplotlib chart, saves it to /charts/, and prints a console report."""
    
    if df.empty:
        print(f"WARNING: Cannot generate {chart_type} '{title}'. DataFrame is empty.", file=sys.stderr)
        return
        
    # Prepare data for plot types that use the DataFrame index (Pie, Bar, Line)
    if chart_type in ['bar', 'barh', 'line', 'pie'] and df.shape[1] >= 2:
        plot_data = df.set_index(df.columns[0])[df.columns[1]]
    else:
        plot_data = df 
        
    plt.figure(figsize=(12, 7))

    # Plotting logic based on chart type
    if chart_type == 'pie':
        plot_data.plot.pie(autopct='%1.1f%%', startangle=90, legend=False, **kwargs)
        plt.ylabel('')
    elif chart_type == 'bar':
        plot_data.plot.bar(**kwargs)
        plt.xticks(rotation=45, ha='right')
    elif chart_type == 'barh':
        plot_data.plot.barh(**kwargs)
        plt.gca().invert_yaxis()
    elif chart_type == 'line':
        plot_data.plot.line(marker='o', **kwargs)
    elif chart_type == 'histogram':
        df[x_col].plot.hist(bins=15, edgecolor='black', alpha=0.7, **kwargs)
    elif chart_type == 'scatter':
        plt.scatter(df[x_col], df[y_col], **kwargs)
        if 'c' in kwargs: 
             plt.legend(*plt.gca().get_legend_handles_labels(), title='Alignment' if 'alignment' in df.columns else df.columns[-1])
    
    # General chart aesthetics
    plt.title(title)
    if x_label and chart_type not in ['pie']:
        plt.xlabel(x_label)
    if y_label and chart_type not in ['pie']:
        plt.ylabel(y_label)
        
    plt.grid(axis='y', linestyle='--', alpha=0.7)
    plt.tight_layout()
    
    save_path = os.path.join('charts', filename)
    plt.savefig(save_path)
    plt.close()

    print(f"Chart Type: {chart_type.upper()}. Saved to: {save_path}")
    print(f"Shows: {title}")


# --- TASK 1: MATPLOTLIB VISUALIZATIONS (60 POINTS) ---
# **SQL QUERIES ARE MOVED HERE TO DEFINE THEM BEFORE MAIN EXECUTION**

# 1. Pie Chart: Alignment Distribution (2 JOINs)
query_pie = """
SELECT al.alignment, COUNT(s.id) AS hero_count
FROM "superhero"."superhero" s 
JOIN "superhero"."alignment" al ON s.alignment_id = al.id 
GROUP BY al.alignment;
"""

# 2. Bar Chart: Top 10 Publishers by Total Powers (3 JOINs)
query_bar = """
SELECT p.publisher_name, COUNT(hp.power_id) AS total_powers_count
FROM "superhero"."publisher" p 
JOIN "superhero"."superhero" s ON s.publisher_id = p.id 
JOIN "superhero"."hero_power" hp ON s.id = hp.hero_id 
GROUP BY p.publisher_name
HAVING COUNT(hp.power_id) > 50
ORDER BY total_powers_count DESC
LIMIT 10;
"""

# 3. Horizontal Bar Chart: Average Combat Rating by Race (3 JOINs)
query_hbar = """
SELECT r.race, AVG(ha.attribute_value)::numeric(8,2) AS avg_combat_score
FROM "superhero"."superhero" s 
JOIN "superhero"."hero_attribute" ha ON s.id = ha.hero_id 
JOIN "superhero"."attribute" a ON ha.attribute_id = a.id 
JOIN "superhero"."race" r ON s.race_id = r.id 
WHERE a.attribute_name = 'Combat' AND r.race != '-' AND r.race != 'N/A'
GROUP BY r.race
HAVING COUNT(s.id) > 5
ORDER BY avg_combat_score DESC
LIMIT 10;
"""

# 4. Line Chart: Comparison of Avg Strength and Intelligence by Publisher (3 JOINs)
query_line = """
SELECT p.publisher_name, a.attribute_name, AVG(ha.attribute_value) AS avg_attribute_value
FROM "superhero"."superhero" s 
JOIN "superhero"."publisher" p ON s.publisher_id = p.id 
JOIN "superhero"."hero_attribute" ha ON s.id = ha.hero_id 
JOIN "superhero"."attribute" a ON ha.attribute_id = a.id 
WHERE p.publisher_name IN ('Marvel Comics', 'DC Comics')
  AND a.attribute_name IN ('Strength', 'Intelligence')
GROUP BY p.publisher_name, a.attribute_name
ORDER BY p.publisher_name, a.attribute_name;
"""

# 5. Histogram: Height Distribution for Marvel Comics Heroes (2 JOINs)
query_hist = """
SELECT s.height_cm
FROM "superhero"."superhero" s 
JOIN "superhero"."publisher" p ON s.publisher_id = p.id 
WHERE p.publisher_name = 'Marvel Comics' AND s.height_cm IS NOT NULL AND s.height_cm > 0;
"""

# 6. Scatter Plot: Durability vs Power for 'Good' Alignment Heroes (3 JOINs + CTE)
query_scatter = """
SELECT T1.superhero_name,
       MAX(CASE WHEN T2.attribute_name = 'Durability' THEN T3.attribute_value END) AS Durability,
       MAX(CASE WHEN T2.attribute_name = 'Power' THEN T3.attribute_value END) AS Power,
       T4.alignment
FROM "superhero"."superhero" T1 
JOIN "superhero"."attribute" T2 ON 1=1
JOIN "superhero"."hero_attribute" T3 ON T1.id = T3.hero_id AND T2.id = T3.attribute_id 
JOIN "superhero"."alignment" T4 ON T1.alignment_id = T4.id 
WHERE T4.alignment = 'Good' AND T2.attribute_name IN ('Durability', 'Power')
GROUP BY T1.superhero_name, T4.alignment;
"""


# --- TASK 2: PLOTLY TIME SLIDER (15 POINTS) ---

def generate_plotly_data():
    """Fetches data and generates a 'year_of_debut' column for Plotly animation. (4 JOINs)"""
    query_plotly = """
    SELECT s.id, s.superhero_name, 
           MAX(CASE WHEN a.attribute_name = 'Intelligence' THEN ha.attribute_value END) AS intelligence,
           MAX(CASE WHEN a.attribute_name = 'Strength' THEN ha.attribute_value END) AS strength,
           p.publisher_name,
           al.alignment
    FROM "superhero"."superhero" s 
    JOIN "superhero"."hero_attribute" ha ON s.id = ha.hero_id 
    JOIN "superhero"."attribute" a ON ha.attribute_id = a.id 
    LEFT JOIN "superhero"."publisher" p ON s.publisher_id = p.id 
    LEFT JOIN "superhero"."alignment" al ON s.alignment_id = al.id 
    GROUP BY s.id, s.superhero_name, p.publisher_name, al.alignment
    HAVING MAX(CASE WHEN a.attribute_name = 'Intelligence' THEN ha.attribute_value END) IS NOT NULL
       AND MAX(CASE WHEN a.attribute_name = 'Strength' THEN ha.attribute_value END) IS NOT NULL
       AND p.publisher_name IN ('Marvel Comics', 'DC Comics');
    """
    df_plotly = run_query(query_plotly, "Plotly Data Fetch")
    
    # Generate mock time column (as data lacks one)
    min_year = 1990
    max_year = 2020
    year_range = max_year - min_year
    
    # Generate a year based on ID, mapping them to the range.
    df_plotly['year_of_debut'] = df_plotly['id'].apply(lambda x: min_year + (x % (year_range + 1)))
    
    # Ensure it's an integer
    df_plotly['year_of_debut'] = df_plotly['year_of_debut'].astype(int) 
    
    # CRUCIAL FIX 1: Sort the DataFrame by the year column to ensure Plotly animation runs chronologically
    df_plotly = df_plotly.sort_values(by='year_of_debut', ascending=True).reset_index(drop=True)
    
    return df_plotly

def generate_plotly_slider(df_plotly):
    """Generates the Plotly scatter plot with a time slider and saves it to HTML."""
    if df_plotly.empty:
        print("\nNo data to build the interactive Plotly graph.")
        return
        
    print("\n--- Task 2: Interactive Plotly Graph (HTML Save) ---")
    
    # CRUCIAL FIX 2: Create year_of_debut_str ONLY here, as it's needed only for Plotly animation grouping/labels
    df_plotly['year_of_debut_str'] = df_plotly['year_of_debut'].astype(str)
    
    fig = px.scatter(df_plotly, 
                     x="intelligence", 
                     y="strength",
                     animation_frame="year_of_debut_str", # Используем строковое представление для правильного порядка
                     size="intelligence", 
                     color="publisher_name",
                     hover_name="superhero_name",
                     log_x=False, size_max=40,
                     title="Interactive Analysis: Hero Intelligence vs. Strength Over Time (Mock Data)")
    
    html_path = os.path.join('charts', 'plotly_timeslider_interactive.html')

    # Guaranteed save to HTML
    fig.write_html(html_path)
    print(f"Interactive chart saved to: {html_path}.")
    print("For demonstration, open this HTML file manually in your browser.")
    
# --- NEW: EXPORT PLOTLY DATA TO SEPARATE EXCEL FILE (User Request) ---

def export_plotly_data_to_excel(df, filename):
    """Exports the Plotly data (with mock year) to a separate Excel file."""
    if df.empty:
        print(f"WARNING: Cannot export {filename}. DataFrame is empty.", file=sys.stderr)
        return
        
    full_path = os.path.join('exports', filename)
    
    # FIX for KeyError: Ensure we drop the 'year_of_debut_str' column only if it exists
    # Since we moved its creation to generate_plotly_slider, it should not exist here unless
    # called multiple times, but let's be explicit and remove the dropping attempt. 
    # We will rename the DataFrame to avoid confusion with the version that has '_str'.
    df_export = df.rename(columns={'year_of_debut': 'Year_of_Debut'})
    
    with pd.ExcelWriter(full_path, engine='openpyxl') as writer:
        df_export.to_excel(writer, sheet_name="Plotly Data", index=False)
        
        # Добавляем базовое форматирование для читаемости
        ws = writer.sheets["Plotly Data"]
        ws.freeze_panes = "B2"
        ws.auto_filter.ref = ws.dimensions
        
    print(f"Plotly data successfully exported to: {full_path}")


# --- TASK 3: EXPORT TO EXCEL WITH FORMATTING (25 POINTS) ---

def export_to_excel_final(dataframes_dict, filename="superheroes_report.xlsx"):
    """Exports multiple DataFrames to a single Excel file with complex formatting."""
    from openpyxl.styles import Font, Color 
    
    full_path = os.path.join('exports', filename)
    total_rows = 0
    sheet_count = 0
    
    print(f"\n--- Excel Export Report ---")

    # 1. Write DataFrames to file
    with pd.ExcelWriter(full_path, engine='openpyxl') as writer:
        for sheet_name, df in dataframes_dict.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False, header=True)
            total_rows += len(df)
            sheet_count += 1
            
            # Get the worksheet for formatting
            ws = writer.sheets[sheet_name]
            
            # --- Required Formatting Elements ---
            
            # 1. Freeze Rows/Columns for Headers: Freeze row 1 and column A (B2 pane)
            if ws.max_row > 1 and ws.max_column > 1:
                ws.freeze_panes = "B2"

            # 2. Filters on all columns (Auto Filter)
            if ws.max_row > 1 and ws.max_column > 0:
                ws.auto_filter.ref = ws.dimensions
            
            # Identify numeric columns for conditional formatting
            numeric_cols = [i for i, col in enumerate(df.columns) if pd.api.types.is_numeric_dtype(df[col]) or pd.api.types.is_integer_dtype(df[col])]
            
            for col_index_df in numeric_cols:
                col_letter = chr(ord('A') + col_index_df) 
                data_range = f"{col_letter}2:{col_letter}{ws.max_row}"
                
                if ws.max_row > 1:
                    # 3. Gradient Fill (Color Scale Rule)
                    rule_scale = ColorScaleRule(start_type="min", start_color=Color(rgb="FFAA0000"), 
                                                mid_type="percentile", mid_value=50, mid_color=Color(rgb="FFFFFF00"),
                                                end_type="max", end_color=Color(rgb="FF00AA00")) 
                    ws.conditional_formatting.add(data_range, rule_scale)

                    # 4. Conditional Formatting (Highlighting min/max)
                    min_rule = FormulaRule(
                        formula=[f"={col_letter}2=MIN(${col_letter}$2:${col_letter}${ws.max_row})"],
                        font=Font(bold=True, color=Color(rgb="FFC70039"))
                    )
                    max_rule = FormulaRule(
                        formula=[f"={col_letter}2=MAX(${col_letter}$2:${col_letter}${ws.max_row})"],
                        font=Font(bold=True, color=Color(rgb="FF006100"))
                    )
                    ws.conditional_formatting.add(data_range, min_rule)
                    ws.conditional_formatting.add(data_range, max_rule)
                    
    print(f"Created file: {filename}, {sheet_count} sheets, {total_rows} rows.")
    return total_rows, sheet_count

# SQL queries for Excel export (complex reports)
query_excel_1 = """
-- Top 100 heroes by Total Attributes, showing Publisher/Race/Alignment (4 JOINs)
SELECT s.superhero_name, p.publisher_name, r.race, al.alignment,
       COALESCE(SUM(ha.attribute_value), 0) AS total_attributes
FROM "superhero"."superhero" s 
LEFT JOIN "superhero"."hero_attribute" ha ON s.id = ha.hero_id
LEFT JOIN "superhero"."publisher" p ON s.publisher_id = p.id
LEFT JOIN "superhero"."race" r ON s.race_id = r.id
LEFT JOIN "superhero"."alignment" al ON s.alignment_id = al.id
GROUP BY s.superhero_name, p.publisher_name, r.race, al.alignment
ORDER BY total_attributes DESC
LIMIT 100;
"""

query_excel_2 = """
-- Distribution of Powers by Publisher (3 JOINs)
SELECT sp.power_name, p.publisher_name, COUNT(DISTINCT hp.hero_id) AS heroes_with_power
FROM "superhero"."superpower" sp
JOIN "superhero"."hero_power" hp ON sp.id = hp.power_id
JOIN "superhero"."superhero" s ON hp.hero_id = s.id
JOIN "superhero"."publisher" p ON s.publisher_id = p.id
GROUP BY sp.power_name, p.publisher_name
ORDER BY sp.power_name, heroes_with_power DESC;
"""

# --- MAIN EXECUTION BLOCK ---

if __name__ == '__main__':
    
    # --- TASK 1: EXECUTE ALL MATPLOTLIB VISUALIZATIONS ---
    
    # 1. Pie Chart
    df_pie = run_query(query_pie, "Distribution of Heroes by Alignment")
    generate_chart(df_pie, 'pie', "Distribution of Heroes by Moral Alignment", "pie_alignment.png")

    # 2. Bar Chart
    df_bar = run_query(query_bar, "Top 10 Publishers by Total Assigned Superpowers")
    generate_chart(df_bar, 'bar', "Top 10 Publishers by Total Assigned Superpowers", "bar_publisher_powers.png", 
                   x_label="Publisher", y_label="Total Power Count")

    # 3. Horizontal Bar Chart
    df_hbar = run_query(query_hbar, "Average Combat Rating by Race (Top 10)")
    generate_chart(df_hbar, 'barh', "Average Combat Rating by Race (Top 10)", "hbar_avg_combat_race.png", 
                   x_label="Average Combat Score", y_label="Race")

    # 4. Line Chart
    df_line = run_query(query_line, "Comparison of Avg Intelligence and Strength: Marvel vs DC")
    
    # Manual pivoting for Matplotlib line chart
    df_line_pivot = df_line.pivot(index='publisher_name', columns='attribute_name', values='avg_attribute_value').reset_index()
    
    plt.figure(figsize=(10, 6))
    plt.plot(df_line_pivot['publisher_name'], df_line_pivot['Intelligence'], marker='o', label='Intelligence', color='blue')
    plt.plot(df_line_pivot['publisher_name'], df_line_pivot['Strength'], marker='x', label='Strength', color='red')
    plt.title("Comparison of Avg Intelligence and Strength: Marvel vs DC")
    plt.xlabel("Publisher")
    plt.ylabel("Average Attribute Value")
    plt.legend(title="Attribute")
    plt.grid(axis='y', linestyle='--', alpha=0.7)
    plt.tight_layout()
    plt.savefig(os.path.join('charts', "line_intel_vs_strength_md.png"))
    plt.close()
    print("Chart Type: LINE. Saved to: charts/line_intel_vs_strength_md.png")
    print("Shows: Trend comparison of average strength and intelligence between Marvel and DC.")

    # 5. Histogram
    df_hist = run_query(query_hist, "Height Distribution for Marvel Comics Heroes")
    generate_chart(df_hist, 'histogram', "Height Distribution for Marvel Comics Heroes (cm)", "hist_marvel_height.png", 
                   x_label="Height (cm)", y_label="Frequency (Number of Heroes)", x_col='height_cm')

    # 6. Scatter Plot
    df_scatter = run_query(query_scatter, "Durability vs Power for 'Good' Alignment Heroes")
    generate_chart(df_scatter, 'scatter', "Durability (X) vs Power (Y) for 'Good' Heroes", "scatter_durability_vs_power.png", 
                   x_label="Durability Rating", y_label="Power Rating", 
                   x_col='durability', y_col='power')

    # --- TASK 3: EXECUTE EXPORT TO MAIN EXCEL REPORT ---
    df_excel_1 = run_query(query_excel_1, "Excel Data: Top 100 Heroes by Total Attributes")
    df_excel_2 = run_query(query_excel_2, "Excel Data: Power Distribution by Publisher")

    dataframes_to_export = {
        "Top Heroes Report": df_excel_1,
        "Power X Publisher": df_excel_2
    }
    
    # Saving to the main Excel report
    export_to_excel_final(dataframes_to_export, "superheroes_report.xlsx")

    # --- TASK 2: PLOTLY SLIDER DEMO (Interactive HTML Save) ---
    df_plotly_data = generate_plotly_data()
    
    # 1. Save data to a separate Excel file (New requirement)
    export_plotly_data_to_excel(df_plotly_data, "plotly_slider_data.xlsx")
    
    # 2. Generate and save the interactive HTML chart
    generate_plotly_slider(df_plotly_data) 
    
    print("\n--- Project Assignment #2 Completed ---")
    print("Check the 'charts/' and 'exports/' folders for results.")
